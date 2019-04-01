import openpyxl as xl
import datetime as dt

class ExcelWrapper:
    book = None
    template = None
    data = None
    new_user = 0
    lookup = dict()     # Names -> Row number

    def __init__(self, template=None, debug=False):
        self.debug = debug
        if self.debug:
            global dbg
            import debug_tools as dbg
        try:
            self.book = xl.reader.excel.load_workbook(template)
        except:
            self.book = xl.workbook.Workbook()
            self.book.active.title = 'Template'
            self.book.create_sheet('Data')
        self.template = self.book['Template']
        self.data = self.book['Data']
        self.data.ins = 0
        self.data.outs = 0
        self._make_lookup()

    def save(self, file):
        self.book.remove(self.template)
        self.book.save(file)

    def fill(self, data): self._fill_table(*(self._fill_data(data)))

    def _add_day(self, date):
        table = self.book.copy_worksheet(self.template)
        table.title = date.strftime('%b %d')
        return self.book[table.title]

    def _add_user(self, name, ts, col, row):
        self.data.cell(column=16, row=self.new_user+3).value = name
        self.data.cell(column=17, row=self.new_user+3).value = ts
        self.data.cell(column=18, row=self.new_user+3).value = 'Col: ' + str(col)
        self.data.cell(column=19, row=self.new_user+3).value = 'Row: ' + str(row)
        self.new_user += 1

    def _fill_data(self, data):
        # data is a list of tuples (last_name, first_name, date, time, message_text, ts, in/out_flag).
        # High and low times, and flags to indicate first shift.
        hi = dt.datetime.min
        lo = dt.datetime.max
        in_row = 3
        out_row = 3
        unk_row = 3
        for last, first, date, time, msg, ts, io in data:
            ts = dt.datetime.fromtimestamp(ts)
            if io == 1:   # Punch ins
                col, row = 1, in_row
                in_row += 1
            elif io == -1:   # Punch outs
                col, row = 6, out_row
                out_row += 1
            else:   # Punch unknowns
                col, row = 11, unk_row
                unk_row += 1
            self._fill_datum(col, row, last, first, date, time, msg, ts)
            if ts < lo:
                lo = ts
            elif ts > hi:
                hi = ts
        self.data.ins = in_row - 3
        self.data.outs = out_row - 3
        self.data.unks = unk_row - 3
        return lo.date(), hi.date()

    def _fill_datum(self, col, row, last, first, date, time, msg, ts):
        self.data.cell(column=col, row=row).value = last + ', ' + first
        self.data.cell(column=col+1, row=row).value = date
        self.data.cell(column=col+2, row=row).value = time
        self.data.cell(column=col+3, row=row).value = xl.utils.datetime.to_excel(ts)
        self.data.cell(column=col+4, row=row).value = msg

    def _fill_table(self, start, end):
        in_row = 3
        out_row = 3
        while start <= end:
            self._add_day(start)
            in_row = self._move_data_to_table(4, 1, in_row, start)
            out_row = self._move_data_to_table(5, 6, out_row, start)
            start += dt.timedelta(1)

    def _make_lookup(self):
        for row in self.template.rows:
            if row[2].row == 1 or row[2].value is None:
                # First row is a header, last is empty.
                continue
            name = row[2].value
            if name not in self.lookup:
                self.lookup[name] = row[2].row
        if self.debug: dbg.tree(self.lookup,1,'Names Lookup')

    def _move_data_to_table(self, table_col, data_col, data_row, table_date):
        # col to write to, column to read from, current row read, current date
        empty_cell = set(("", "'", None))
        table = self.book[table_date.strftime('%b %d')]
        data_date = self.data.cell(column=data_col+1, row=data_row).value
        while table_date.strftime('%Y/%m/%d') == data_date:
            name = self.data.cell(column=data_col, row=data_row).value
            ts = self.data.cell(column=data_col+3, row=data_row).value
            try:
                table_row = self.lookup[name]
            except KeyError:
                if self.debug: dbg.bp()
                self._add_user(name, ts, data_col, data_row)
                data_row += 1
                data_date = self.data.cell(column=data_col+1, row=data_row).value
                continue
            cell = table.cell(column=table_col, row=table_row)
            if cell.value in empty_cell:
                cell.comment = xl.comments.Comment(str(data_row),'TimeAuditBot')
                cell.value = ts
            else:
                cell.value = '!!!!!!!!'
                cell.comment.text += ', ' + str(data_row)
            data_row += 1
            data_date = self.data.cell(column=data_col+1, row=data_row).value
        return data_row
